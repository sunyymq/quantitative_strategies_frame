
import os
import calendar
import warnings
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from utility.relate_to_tushare import trade_days, generate_months_ends

try:
    from WindPy import *
except ImportError:
    print('导入Wind接口失败')


# 调整财报数据日期到披露日期
def adjust_months(d_df, orig='Q'):

    if isinstance(d_df.columns[0], str):
        new_cols = [datetime.strptime(col, "%Y-%m-%d") for col in d_df.columns]
        d_df.columns = new_cols

    # 原始数据为季度数据
    if orig == 'Q':
        # 删除12月份的数据
        tdc_1 = [col for col in d_df.columns if col.month == 12]
        # 删除非3、6、9、12的月份的数据
        tdc_2 = [col for col in d_df.columns if col.month not in [3, 6, 9, 12]]

        tdc = tdc_1 + tdc_2
        d_df = d_df.drop(tdc, axis=1)

        # 把公告月份调整为实际月份
        new_cols = []
        for col in d_df.columns:
            if col.month == 3:
                new_cols.append(datetime(col.year, 4, 30))
            elif col.month == 6:
                new_cols.append(datetime(col.year, 8, 31))
            elif col.month == 9:
                new_cols.append(datetime(col.year, 10, 31))
            else:
                print(col)

        d_df.columns = new_cols
    # 原始数据为年度数据
    elif orig == 'Y':
        new_cols = [datetime(col.year + 1, 4, 30) for col in d_df.columns]
        d_df.columns = new_cols

    return d_df


# 把一个 截面数据添加到已经有的月度模式存储的文件中
def add_to_panels(dat, panel_path, f_name, freq_in_dat='M'):
    """说明： 把dat依次插入到panel_path的DF中，插入的列名为f_name, 根据dat的类型是DF还是Series可以判断
    是每次插入的数据不同还是每次插入相同的数据。"""

    print(f'开始添加{f_name}数据到目标文件夹')
    panel = os.listdir(panel_path)
    for month_date in panel:
        hased_dat = pd.read_csv(os.path.join(panel_path, month_date), engine='python')
        hased_dat = hased_dat.set_index('Code')

        # 输入数据为 DataFrame, 那么按列插入
        if isinstance(dat, pd.DataFrame):
            mon_str = month_date.split('.')[0]
            if mon_str in dat.columns:
                # 当dat中的columns也是str格式，且日期与panel一样时，直接添加
                hased_dat[f_name] = dat[mon_str]
            else:
                # 否则，当年、月相同，日不同时，需要变成datetime格式而且还有查找
                target = datetime.strptime(mon_str, "%Y-%m-%d")
                # 当dat的columns是datetime格式时
                if isinstance(dat.columns[0], datetime):
                    if freq_in_dat == 'M':
                        finded = None
                        for col in dat.columns:
                            if col.year == target.year and col.month == target.month:
                                finded = col
                                break
                        if finded:
                            hased_dat[f_name] = dat[finded]
                        else:
                            print('{}该期未找到对应数据'.format(mon_str))
                    if freq_in_dat == 'D':
                        if target in dat.columns:
                            hased_dat[f_name] = dat[target]
                        else:
                            print('{}该期未找到对应数据'.format(mon_str))
                else:
                    print('现有格式的还未完善')
                    raise Exception
        # 输入数据为 DataFrame, 那么按列插入
        elif isinstance(dat, pd.Series):
            hased_dat[f_name] = dat[hased_dat.index]

        try:
            hased_dat = hased_dat.reset_index('Code')
        except Exception as e:
            print('debug')

        if 'No' in hased_dat.columns:
            del hased_dat['No']
        hased_dat.index.name = 'No'
        hased_dat.to_csv(os.path.join(panel_path, month_date), encoding='gbk')

    print('完毕！')


# 根据 months_end 的列拓展 d_df。
# 我从数据库中下载的基本面数据是调整过日期的，但是调整到的是月度的最后一个自然日，要改成月度最后一个交易日，
# 同时还有复制到其他月份中。
def append_df(d_df, target_feq='M', fill_type='preceding'):
    '''
    fill_type ：若整列和为0的填充方式，preceding表示使用前值填充,  empty表示不填充
    '''

    tds = trade_days()

    # 得到月末日期列表
    months_end = generate_months_ends()

    # 赵到对应的月末日期列表，可能年月同日不同的情况
    new_col = []
    for col in d_df.columns:
        for me in months_end:
            if col.year == me.year and col.month == me.month:
                new_col.append(me)
    # 改变月末日期
    d_df.columns = new_col

    if target_feq.upper() == 'M':
        # 设一个日期全的，单值为空的df
        res = pd.DataFrame(index=d_df.index, columns=months_end)
        # 给定日期赋值
        res[d_df.columns] = d_df

    elif target_feq.upper() == 'D':
        new_columns = [d for d in tds if d >= d_df.columns[0]]
        res = pd.DataFrame(index=d_df.index, columns=new_columns)
        # 给定日期赋值
        res[d_df.columns] = d_df

    elif target_feq.upper() == 'W':
        week_ends = trade_days('w')
        # 因为 月末交易日数据（A） 与 一周交易日数据最后一天(B) 不是一一对应也不是B包含A的关系，所以要做一个A与相对应的B的映射
        res = pd.DataFrame(index=d_df.index, columns=week_ends)
        selected_cols = []
        for col, se in d_df.iteritems():
            delta = [we - col for we in week_ends if (we - col).days >= 0]
            selected_cols.append(col + np.min(delta))
        # 给定日期赋值
        res[selected_cols] = d_df

    # 首期赋值为None
    if res.iloc[:, 0].sum() == 0:
        res.iloc[:, 0] = np.nan

    # 若当列为空，则当列数值与前列相同
    if fill_type == 'preceding':
        res_ar = np.array(res)
        [h, l] = res_ar.shape
        for i in range(1, l):
            for j in range(0, h):
                if np.isnan(res_ar[j, i]) and not np.isnan(res_ar[j, i-1]):
                    res_ar[j, i] = res_ar[j, i-1]

        res_df = pd.DataFrame(data=res_ar, index=res.index, columns=res.columns)

    # 删除nan
    res_df.dropna(axis=1, how='all', inplace=True)
    res_df.dropna(axis=0, how='all', inplace=True)
    res_df.sum()

    return res_df


# 把一个panel中的因子重新命名
def rename_factor_in_panels(panel_path, old_name, new_name):
    panel = os.listdir(panel_path)
    for month_date in panel:
        dat_df = pd.read_csv(os.path.join(panel_path, month_date), engine='python')
        if old_name in dat_df.columns:
            dat_df.set_index(dat_df.columns[0], inplace=True)
            dat_df.rename({old_name: new_name}, axis=1, inplace=True)
            dat_df.to_csv(os.path.join(panel_path, month_date), encoding='gbk')

    print('名称修改完毕')


# 针对 dcit的key是日期，value是DataFarame 存储格式的变量的存储和读取
def wr_excel(path, dat_dict=None, w_or_r='w'):
    if w_or_r == 'r':

        wb = load_workbook(filename=path)
        print(wb.sheetnames)
        sheets = wb.sheetnames

        res_dict = {}
        for i in range(len(sheets)):
            names = sheets[i]
            sheet = wb[sheets[i]]
            print(sheets[i])
            dat_df = pd.DataFrame(sheet.values)
            dat_df.columns = dat_df.loc[0, :]
            dat_df.drop(dat_df.index[0], inplace=True)
            dat_df.set_index(dat_df.columns[0], inplace=True)
            dat_df.index = pd.to_datetime(dat_df.index)
            dat_df = dat_df.sort_index()
            dat_df.dropna(inplace=True)
            res_dict.update({names: dat_df})

        return res_dict

    if w_or_r == 'w':
        for key, dat_df in dat_dict.items():
            if isinstance(key, datetime):
                key = key.strftime("%Y-%m-%d")
            dat_df.to_excel(path, sheet_name=key, encoding='gbk')   # engine='openpyxl'
        return 0






if __name__ == "__main__":
    panel_path = r'D:\Database_Stock\因子预处理模块\因子（已预处理）'
    # old_name = 'mgmt_ben_top3m'.upper()
    # new_name = 'Mgmt_ben_top3m'
    # rename_factor_in_panels(panel_path, old_name, new_name)



